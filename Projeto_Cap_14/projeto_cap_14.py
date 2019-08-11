import openpyxl, os, csv

for excelFile in os.listdir('.'):
    if not excelFile.endswith('.xlsx'):# Ignora os arquivos que não sejam xlsx 
        continue 
    
    nameFileExcel = os.path.basename(excelFile)# Captura o nome do arquivo
    
    wb = openpyxl.load_workbook(nameFileExcel) # Carrega o objeto workbook.
    
    for sheetName in wb: # Percorre todas as planilhas do workbook em um loop.
        sheet = wb[sheetName.title] # Captura apenas o nome da Planilha
        
        name = nameFileExcel[:len(nameFileExcel) - 5] # Retira do nome do arquivo a extensão .xlsx
        
        fileName = name + '_' + sheet.title # Cria o nome do arquivo CSV a partir do nome do arquivo Excel e do título da planilha.
        csvFileObj = open(fileName+'.csv','w', newline = '')
        csvWriter = csv.writer(csvFileObj)# Cria o objeto csv.writer para esse arquivo CSV.

        rowData = []
        for rowNum in range(1, sheet.max_row + 1): # Percorre todas as linhas da planilha em um loop.
            row = []
            for colNum in range(1, sheet.max_column + 1):# Percorre todas as células da linha em um loop.
                cel = (sheet.cell(row = rowNum, column = colNum).value)
                row.append(cel)# Adiciona todas as células a essa lista
            rowData.append(row)# Adiciona os dados de cada célula em rowData.

        for row in rowData:# Grava a lista rowData no arquivo CSV.
            csvWriter.writerow(row)
            
        csvFileObj.close()
        
        print('ARQUIVO '+ fileName + '.csv CRIADO COM SUCESSO' )